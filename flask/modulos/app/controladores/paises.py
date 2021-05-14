import os
from flask.json import load
from openpyxl import Workbook
from openpyxl import load_workbook
from flask import request, jsonify
from timeit import default_timer
from app import app,mongo

ROOT_PATH = os.environ.get('ROOT_PATH')
#crear el archivo excel
#variable= "./TiemposFlask.xlsx"
#wb2 = Workbook()
#wb2.save(variable)

variable2= "./TiemposFlask.xlsx"
wb2 = load_workbook(variable2)
ws= wb2.active

@app.route('/paises/tiempo')
def tiempo_ejecutar():  
    inicio3 = default_timer()
    con3= consulta1()
    fin3 = default_timer()
    time1=fin3-inicio3
    ws = wb2['Sheet']
    ws['A1']= 'flask_consultas'
    ws['A2']= 'consulta1' 
    ws['B1']= 'flask_tiempos'     
    ws['B2']=time1
     

    inicio1 = default_timer()
    con1= consulta2()
    fin1 = default_timer()    
    time2=fin1-inicio1  
    ws['A3']= 'consulta2'      
    ws['B3']=time2    


    inicio2 = default_timer()
    con2= consulta3()
    fin2 = default_timer()
    time3=fin2-inicio2  
    ws['A4']= 'consulta3'      
    ws['B4']=time3
 
    wb2.save(variable2) 
    wb2.close()

    return jsonify({'flask_consultasPaises':["consulta1","consulta2","consulta3"],'tiemposcalculo_flask':[time1,time2,time3]})




@app.route('/lista')
def consulta1():
        data = mongo.db.paises.aggregate([{'$match':{'$and':[{"currency.0":"USD"},{"region":"Americas"}]}},
        {'$count':"Total1"}])
        lista = list(data)
        if lista == []:
            mensaje = "salio mal"
        else:
            mensaje ="todo correcto"            
        return jsonify({'mesaje':mensaje})

@app.route('/lista2')
def consulta2():
    if request.method == 'GET':
        data = mongo.db.paises.find({ '$or': [{"languages.eng":"English"},{"languages.eng":"Spanish"}] })
        lista = list(data)        
        if data == None:
            data = []
        return True

@app.route('/lista3')
def consulta3():
    if request.method == 'GET':
        data = mongo.db.paises.aggregate([{'$match':{"currency.0":"USD"}},{'$count':"Total"}])
        lista = list(data)
        if data == None:
            data = []   
        return True
