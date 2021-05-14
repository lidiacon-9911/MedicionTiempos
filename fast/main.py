from fastapi import FastAPI, Request
from fastapi.middleware.cors import CORSMiddleware
from bson.objectid import ObjectId
from pymongo import  MongoClient
from pymongo import collection
from openpyxl import load_workbook
from openpyxl import Workbook
from timeit import default_timer

cluster = MongoClient("mongodb+srv://flask:flask123@cluster0.67ou4.mongodb.net/paises?retryWrites=true&w=majority")
db = cluster["paises"]
collection = db["paises"]

#crear el archivo excel
#variable= "./TiemposFast.xlsx"
#wb2 = Workbook()
#wb2.save(variable)

variable= "C:/Users/Lenovo/Documents/lIDIA/Tendecnias Actuales/VisualCode/MedicionTiempos/flask/TiemposFlask.xlsx"
wb2 = load_workbook(variable)
ws= wb2.active

app = FastAPI()
@app.get("/")
def listar():    
    return ({'hola':'Consulta de tiempos de ejecucion'})

@app.get('/tiempos_calculos')
def tiempo():
    inicio1 = default_timer()
    con1 = lista_paises1()
    fin1 = default_timer()
    time1 = fin1 - inicio1
    ws = wb2['Sheet']
    ws['C1']= 'fast_consultas'
    ws['C2']= 'consulta1FAST' 
    ws['D1']= 'fast_tiempos'     
    ws['D2']=time1
  

    inicio2 = default_timer()
    con2 = lista_paises2()
    fin2 = default_timer()
    time2 = fin2 - inicio2
  
    ws['C3']= 'consulta2FAST'      
    ws['D3']=time2
  

    inicio3 = default_timer()
    con3 = lista_paises3()
    fin3 = default_timer()
    time3 = fin3 - inicio3  
    ws['C4']= 'consulta3FAST2'      
    ws['D4']=time3
    wb2.save(variable)
    wb2.close()

    return({'fastapi_consulta':["consulta1FAST", "consulta2FAST", "consultaFAST3"],'fastapi_tiempos':[time1, time2, time3]})

@app.get('/listar-paises1')
def lista_paises1():
    data =  collection.find({ '$or': [{"languages.eng":"English"},{"languages.eng":"Spanish"}] })
    return ({'mensaje':'correcto'})   

@app.get('/listar-paises2')
def lista_paises2():
    data2 =  collection.aggregate([{'$match':{"currency.0":"USD"}},{'$count':"Total"}])
    return ({'mensaje':'correcto'})

@app.get('/listar-paises3')
def lista_paises3():
    data3 =  collection.aggregate([{'$match':{'$and':[{"currency.0":"USD"},{"region":"Americas"}]}},
{'$count':"Total1"}])
    return ({'mensaje':'correcto'})    